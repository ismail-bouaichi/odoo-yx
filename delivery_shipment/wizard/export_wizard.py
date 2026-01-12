# -*- coding: utf-8 -*-
from odoo import models, fields, api, _
from odoo.exceptions import UserError
import base64
import io

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class DeliveryShipmentExportWizard(models.TransientModel):
    _name = 'delivery.shipment.export.wizard'
    _description = 'Export Shipments to Excel for Barid'

    shipment_ids = fields.Many2many(
        'delivery.shipment',
        string='Shipments',
        default=lambda self: self._default_shipment_ids(),
    )
    
    @api.model
    def _default_shipment_ids(self):
        """Get shipments from context (active_ids)."""
        if self.env.context.get('active_model') == 'delivery.shipment':
            return self.env.context.get('active_ids', [])
        return []
    
    excel_file = fields.Binary(
        string='Excel File',
        readonly=True,
    )
    
    filename = fields.Char(
        string='Filename',
        readonly=True,
    )
    
    state = fields.Selection([
        ('draft', 'Draft'),
        ('done', 'Done'),
    ], default='draft')

    def action_export(self):
        """Export shipments to Excel with barcodes."""
        self.ensure_one()
        
        if not OPENPYXL_AVAILABLE:
            raise UserError(_("Please install openpyxl library: pip install openpyxl"))
        
        if not self.shipment_ids:
            raise UserError(_("No shipments selected for export."))
        
        # Check all shipments have packages with GAB
        missing_gab = self.shipment_ids.filtered(lambda s: not s.package_ids)
        if missing_gab:
            raise UserError(_(
                "The following shipments are missing GAB barcodes:\n%s\n\n"
                "Please generate barcodes first."
            ) % '\n'.join(missing_gab.mapped('name')))
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Barid Export"
        
        # Headers - matching Amana format: GAB, ETOILE, CAB1
        headers = ['GAB', 'ETOILE', 'CAB1', 'Nom', 'Prénom', 'Code Postal', 'Ville', 'Adresse', 'MS Destinataire']
        
        # Style for headers
        header_font = Font(bold=True, size=11)
        header_alignment = Alignment(horizontal='center', vertical='center')
        header_fill = PatternFill(start_color='93C572', end_color='93C572', fill_type='solid')  # Pistachio green
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Barcode font style (C39HrP24DhTt Code 39 font)
        barcode_font = Font(name='C39HrP24DhTt', size=24)
        barcode_alignment = Alignment(horizontal='center', vertical='center')
        
        # Write headers
        ws.row_dimensions[1].height = 30  # Header row height
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
            cell.fill = header_fill
        
        # Set column widths
        ws.column_dimensions['A'].width = 35  # GAB barcode
        ws.column_dimensions['B'].width = 8   # ETOILE (just *)
        ws.column_dimensions['C'].width = 35  # CAB1 barcode
        ws.column_dimensions['D'].width = 15  # Nom
        ws.column_dimensions['E'].width = 15  # Prénom
        ws.column_dimensions['F'].width = 12  # Code Postal
        ws.column_dimensions['G'].width = 20  # Ville
        ws.column_dimensions['H'].width = 40  # Adresse
        ws.column_dimensions['I'].width = 20  # MS Destinataire
        
        # Write data - one row per package (not per shipment)
        row_idx = 2
        for shipment in self.shipment_ids:
            partner = shipment.partner_id
            
            # Partner info - split name into nom/prenom
            full_name = partner.name or ''
            name_parts = full_name.split(' ', 1)
            prenom = name_parts[0] if name_parts else ''
            nom = name_parts[1] if len(name_parts) > 1 else ''
            
            # Address
            address_parts = [
                partner.street or '',
                partner.street2 or '',
            ]
            address = ', '.join(filter(None, address_parts))
            
            # Create one row for each package in the shipment
            for package in shipment.package_ids:
                # Set row height for barcode font
                ws.row_dimensions[row_idx].height = 60
                
                # GAB - raw barcode value with barcode font
                gab_cell = ws.cell(row=row_idx, column=1, value=package.gab)
                gab_cell.font = barcode_font
                gab_cell.alignment = barcode_alignment
                gab_cell.border = thin_border
                
                # ETOILE - just an asterisk * (no barcode font)
                etoile_cell = ws.cell(row=row_idx, column=2, value='*')
                etoile_cell.alignment = barcode_alignment
                etoile_cell.border = thin_border
                
                # CAB1 - Excel formula =CONCATENATE(B,A,B) for *GAB*
                cab1_cell = ws.cell(row=row_idx, column=3, value=f'=CONCATENATE(B{row_idx},A{row_idx},B{row_idx})')
                cab1_cell.font = barcode_font
                cab1_cell.alignment = barcode_alignment
                cab1_cell.border = thin_border
                
                # Center alignment for all data cells
                cell_alignment = Alignment(horizontal='center', vertical='center')
                
                nom_cell = ws.cell(row=row_idx, column=4, value=nom)
                nom_cell.border = thin_border
                nom_cell.alignment = cell_alignment
                
                prenom_cell = ws.cell(row=row_idx, column=5, value=prenom)
                prenom_cell.border = thin_border
                prenom_cell.alignment = cell_alignment
                
                zip_cell = ws.cell(row=row_idx, column=6, value=partner.zip or '')
                zip_cell.border = thin_border
                zip_cell.alignment = cell_alignment
                
                city_cell = ws.cell(row=row_idx, column=7, value=partner.city or '')
                city_cell.border = thin_border
                city_cell.alignment = cell_alignment
                
                address_cell = ws.cell(row=row_idx, column=8, value=address)
                address_cell.border = thin_border
                address_cell.alignment = cell_alignment
                
                # MS Destinataire
                ms_cell = ws.cell(row=row_idx, column=9, value=shipment.ms_destinataire or '')
                ms_cell.border = thin_border
                ms_cell.alignment = cell_alignment
                
                row_idx += 1
        
        # Save to buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Update wizard
        self.write({
            'excel_file': base64.b64encode(output.read()),
            'filename': f'barid_export_{fields.Date.today()}.xlsx',
            'state': 'done',
        })
        
        # Return wizard to download file
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }
